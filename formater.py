import math
import pandas as pd
from io import BytesIO
import base64
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, numbers
import numpy as np
import warnings

def custom_warning_handler(message, category, filename, lineno, file=None, line=None):
    if "Cannot parse header or footer" in str(message):
        return
    warnings.showwarning(message, category, filename, lineno, file, line)

warnings.showwarning = custom_warning_handler

class ExcelDecoder:
    @staticmethod
    def decode_content(contentBytes, skiprows=0, skipfooter=0):
        decoded = base64.b64decode(contentBytes)
        excel = BytesIO(decoded)
        df = pd.read_excel(excel, skiprows=skiprows, skipfooter=skipfooter, engine='openpyxl')
        return df

class DataProcessor:
    @staticmethod
    def filter_and_group_data(df):
        df = df[df['Live Check Amount'] > 0]
        grouped_data = df.groupby('Client').agg(
            number_of_live_checks=pd.NamedAgg(column='Live Check Amount', aggfunc='count'),
            check_totals=pd.NamedAgg(column='Live Check Amount', aggfunc='sum')
        ).reset_index()
        return grouped_data

    @staticmethod
    def add_totals_row(grouped_data):
        grouped_data.loc[len(grouped_data)] = {
            'Client': 'Totals',
            'number_of_live_checks': grouped_data['number_of_live_checks'].sum(),
            'check_totals': grouped_data['check_totals'].sum()
        }
        return grouped_data

class ExcelFormatter:
    @staticmethod
    def format_worksheet(grouped_data):
        wb = Workbook()
        ws = wb.active
        for r in dataframe_to_rows(grouped_data, index=False, header=True):
            ws.append(r)
        ExcelFormatter._apply_styles(ws)
        return wb

    @staticmethod
    def _apply_styles(ws):
        blue_fill = PatternFill(start_color="94DCF8", end_color="94DCF8", fill_type="solid")
        for cell in ws[1]:
            cell.fill = blue_fill
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        for cell in ws['C']:
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        last_row = ws[len(ws['A'])]
        for cell in last_row:
            cell.fill = blue_fill

class ExcelSaver:
    @staticmethod
    def save_workbook(wb):
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

def formatExcel(contentBytes):
    df = ExcelDecoder.decode_content(contentBytes, skiprows=5, skipfooter=4)
    grouped_data = DataProcessor.filter_and_group_data(df)
    grouped_data = DataProcessor.add_totals_row(grouped_data)
    wb = ExcelFormatter.format_worksheet(grouped_data)
    return ExcelSaver.save_workbook(wb)

def formatFromJson(content):
    df = pd.DataFrame(content)
    df['Live Check Amount'] = df['Live Check Amount'].replace('', np.nan).fillna(0).astype(float)
    grouped_data = DataProcessor.filter_and_group_data(df)
    grouped_data = DataProcessor.add_totals_row(grouped_data)
    return grouped_data.to_dict(orient='records')

def discrepancies_report_ssn(contentBytes, path):
    df = ExcelDecoder.decode_content(contentBytes)
    if "aetna" in path.lower():
        dfs = split_dataframe(df)
       
        for df in dfs:
            ssn_columns = [col for col in df.columns if isinstance(col, str) and pd.notna(col) and 'ssn' in col.lower()]
            for ssn_column in ssn_columns:
                df[ssn_column] = df[ssn_column].apply(remove_leading_zero)
        columns_to_keep = {
        'EE SSN': 'SSN',
        'SSN': 'SSN',
        'Comments': 'Comments',
        'comments': 'Comments',
        'Notes': 'Notes',
        'notes': 'Notes',
        'Dep SSN':'Dep SSN'
        }
        combined_df = pd.DataFrame()
        
        for df in dfs:
            filtered_df = df[[col for col in df.columns if col in columns_to_keep]].rename(columns=columns_to_keep)
            if 'EE SSN' in filtered_df.columns and 'SSN' in filtered_df.columns:
                filtered_df['SSN'] = filtered_df['EE SSN'].combine_first(filtered_df['SSN'])
                filtered_df = filtered_df.drop(columns=['EE SSN'])
            filtered_df['Carrier'] = 'aetna'
            filtered_df['PEO_ID'] = ''
            combined_df = pd.concat([combined_df, filtered_df], ignore_index=True)
        ssn_col = next((col for col in combined_df.columns if col.lower() in ['ssn', 'full ssn', 'ee ssn']), None)
        if ssn_col:
            ssn_records = combined_df[ssn_col].tolist()
            ssn_records = [int(ssn) if isinstance(ssn, str) and ssn.isdigit() else ssn for ssn in ssn_records]
        return {"ssn":ssn_records,"dep_ssn":ssn_records}

    elif "legal shield" in path.lower():
        ssn_col = next((col for col in df.columns if col.lower() in ['ssn', 'full ssn', 'ee ssn']), None)
        if ssn_col:
            ssn_records = df[ssn_col].tolist()
            ssn_records = [int(ssn) if isinstance(ssn, str) and ssn.isdigit() else ssn for ssn in ssn_records]
        return {"ssn":ssn_records,"dep_ssn":ssn_records}
    elif "empire" in path.lower():
        df.columns = [col.strip() for col in df.columns]
        ssn_col = next((col for col in df.columns if col.lower() in ['ssn', 'full ssn', 'ee ssn']), None)
        if ssn_col:
            ssn_records = df[ssn_col].tolist()
            ssn_records = [int(ssn) if isinstance(ssn, str) and ssn.isdigit() else ssn for ssn in ssn_records]
        return {"ssn":ssn_records,"dep_ssn":ssn_records}

    
def discrepancies_report(contentBytes, path, planTermDetails):
    planTermDetails['EE_SSN'] = planTermDetails['EE_SSN'].apply(remove_leading_zero)
    planTermDetails['DEP_SSN'] = planTermDetails['DEP_SSN'].apply(remove_leading_zero)
    df = ExcelDecoder.decode_content(contentBytes)
    df = df.astype(str)
    if "aetna" in path.lower():
        dfs = split_dataframe(df)
        for df in dfs:
            ssn_columns = [col for col in df.columns if isinstance(col, str) and pd.notna(col) and 'ssn' in col.lower()]
            for ssn_column in ssn_columns:
                df[ssn_column] = df[ssn_column].apply(remove_leading_zero)
        columns_to_keep = {
        'EE SSN': 'SSN',
        'SSN': 'SSN',
        'Comments': 'Comments',
        'comments': 'Comments',
        'Notes': 'Notes',
        'notes': 'Notes',
        'Dep SSN':'Dep SSN'
        }
        combined_df = pd.DataFrame()
        
        for df in dfs:

            filtered_df = df[[col for col in df.columns if col in columns_to_keep]].rename(columns=columns_to_keep)
            if 'EE SSN' in filtered_df.columns and 'SSN' in filtered_df.columns:
                filtered_df['SSN'] = filtered_df['EE SSN'].combine_first(filtered_df['SSN'])
                filtered_df = filtered_df.drop(columns=['EE SSN'])
            filtered_df['Carrier'] = 'aetna'
            filtered_df['PEO_ID'] = ''
            combined_df = pd.concat([combined_df, filtered_df], ignore_index=True)
        combined_df['SSN'] = combined_df['SSN'].apply(remove_leading_zero)
        combined_df['Dep SSN'] = combined_df['Dep SSN'].apply(remove_leading_zero)
        combined_df=find_requirement(combined_df,planTermDetails,"DISCREPANCIES.xlsx", 'Comments', 'SSN', 'Dep SSN')
        combined_df.to_excel("tests.xlsx")
        return save_tables_to_excel([combined_df])

    elif "legal shield" in path.lower():
        ssn_columns = [col for col in df.columns if isinstance(col, str) and pd.notna(col) and 'ee ssn' in col.lower()]
        for ssn_column in ssn_columns:
            df[ssn_column] = df[ssn_column].apply(remove_leading_zero)
        df=find_requirement_legalShield(df,planTermDetails)
        return save_tables_to_excel([df])
    elif "empire" in path.lower():
        ssn_columns = [col for col in df.columns if isinstance(col, str) and pd.notna(col) and 'ssn' in col.lower()]
        for ssn_column in ssn_columns:
            df[ssn_column] = df[ssn_column].apply(remove_leading_zero)
        df=find_requirement(df, planTermDetails, "DISCREPANCIES.xlsx", 'HOW TO RESOLVE  (ERROR DESCRIPTION)', 'SSN',dep_ssn_column='SSN',columnsTOKeep=['SSN','Instance', 'HOW TO RESOLVE  (ERROR DESCRIPTION)', 'Found Data', 'key word'])
        return save_tables_to_excel([df])

def remove_leading_zero(ssn):
    if isinstance(ssn, int):
        ssn = str(ssn)
    if pd.notna(ssn):
        if len(ssn) > 9:
            ssn = ssn.lstrip('0')
        elif len(ssn)<9:
            ssn=ssn.zfill(9)
    return ssn


def split_dataframe(df):
    # Lista para almacenar los DataFrames resultantes
    dfs = []
    current_df = []
    in_block = False
    dfsReturn=[]
    for index, row in df.iterrows():
        csa_present = "CSA" in row.values
        name_present = any(name in row.values for name in ["Name", "EE Name", "Dep Name"])
        
        if csa_present and name_present:
            if current_df:
                dfs.append(pd.DataFrame(current_df))
                current_df = []
            in_block = True

        if (row == 'nan').all():
            if current_df:
                dfs.append(pd.DataFrame(current_df))
                current_df = []
            in_block = False

        if in_block:
            current_df.append(row.values)
    for df in dfs:
        new_columns = df.iloc[0]
        dfsReturn.append(pd.DataFrame(df.iloc[1:].values, columns=new_columns)) 
    
    return dfsReturn

def save_tables_to_excel(tables):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        for idx, table in enumerate(tables):
            sheet_name = f"Table_{idx + 1}"
            table.to_excel(writer, sheet_name=sheet_name, index=False)
    output.seek(0)
    return output.getvalue()


def find_requirement_legalShield(df,carrierPlanDetails):
    for index, item in df.iterrows():
        item_ssn = str(item["FULL SSN"])
        carrierPlanDetails['EE SSN'] = carrierPlanDetails['EE SSN'].astype(str)
        resultado = carrierPlanDetails[carrierPlanDetails['EE SSN'] == item_ssn]
          
        field='COVERAGE_END_DATE'
        datos=resultado[field].values
        
        datos_filtrados = datos_filtrados = list(
            {dato for dato in datos if '/' not in str(dato) and not (isinstance(dato, float) and math.isnan(dato))} |
            {dato for dato in datos if '/' in str(dato) and not (isinstance(dato, float) and math.isnan(dato))}
        )
        
        if len(datos_filtrados)==0:
            field='TERMDATE'
            resultado=carrierPlanDetails[carrierPlanDetails['EE SSN'] == item_ssn]
            datos=resultado[field].values
            datos = [
                f"{date.astype('datetime64[D]').item().month}/{date.astype('datetime64[D]').item().day}/{date.astype('datetime64[D]').item().year}"
                if isinstance(date, np.datetime64) else date
                for date in datos if not pd.isna(date)
            ]
            datos_filtrados = list({dato for dato in datos if '/' not in str(dato)} | {dato for dato in datos if '/' in str(dato)})
            datos_joined = ';'.join(map(str, datos_filtrados))
        if resultado.empty:
            datos_joined="User not found"  
        df.at[index, field] = datos_joined
    
    return df


def find_requirement(df, carrierPlanDetails, discrepancies_file, comment_column, ssn_column, dep_ssn_column=None, columnsTOKeep=None):
    discrepancies = pd.read_excel(discrepancies_file)
    df.columns = [col.strip() for col in df.columns]
    df['Found Data'] = ''
    
    for index, item in df.iterrows():
        
        comment = item[comment_column]
        found_keywords = find_keywords(comment, discrepancies)
        
        if found_keywords:
           
            key_word = found_keywords[0]
            item_ssn = item[ssn_column]
            if pd.notna(key_word["Data Base"]):
                df.at[index, 'key word'] = key_word["Data Base"]
                resultado = carrierPlanDetails[carrierPlanDetails['EE_SSN'] == item_ssn]
                if not resultado.empty:
                    df.at[index, 'Found Data'] = filter_and_join_data(resultado, key_word)
                    df.at[index, 'Instance'] = ','.join(resultado["PEO_ID"].unique())
                else:
                    dep_ssn = item[dep_ssn_column]
                    resultado = carrierPlanDetails[carrierPlanDetails["DEP_SSN"] == dep_ssn]
                    if not resultado.empty:
                        df.at[index, 'Found Data'] = filter_and_join_data(resultado, key_word)
                        df.at[index, 'Instance'] = ','.join(resultado["PEO_ID"].unique())
                    else:
                        df.at[index, 'Found Data'] = 'User not found'
            else:
                df.at[index, 'key word'] = 'Invalid field'
                df.at[index, 'Found Data'] = ''
        else:
            df.at[index, 'key word'] = 'There is no keywords'
            df.at[index, 'Found Data'] = ''
    if columnsTOKeep is None:
        df.to_excel("test.xlsx")
        return df
    else:
        return df[columnsTOKeep]

def find_keywords(comment, discrepancies):
    found_keywords = []
    for _, keyword_row in discrepancies.iterrows():
        keyword = str(keyword_row['Key word'])
        if keyword.lower() in comment.lower():
            found_keywords.append(keyword_row.to_dict())
    found_keywords = {item['Data Base']: item for item in found_keywords}.values()
    return list(found_keywords)

def filter_and_join_data(resultado, key_word):
    datos = resultado[key_word["Data Base"]].values
    datos_filtrados = list({dato for dato in datos if '/' not in str(dato)} | {dato for dato in datos if '/' in str(dato)})
    return ';'.join(map(str, datos_filtrados))
